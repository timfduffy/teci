#!/usr/bin/env python3
"""Build Qwen & Gemma official benchmark spreadsheet (models <=35B, Aug 2023 - Aug 2026)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADERS = ["Family", "Generation", "Model", "Params (B)", "Architecture", "Release",
           "Variant / mode", "Benchmark", "Eval config / notes", "Score", "Source"]

qwen_rows = []
gemma_rows = []

def add_gen(rows, family, generation, models, benches, source):
    """models: list of (name, params, arch, release, variant); benches: list of (bench, config, [scores])"""
    for bench, config, scores in benches:
        for (name, params, arch, release, variant), score in zip(models, scores):
            if score is None:
                continue
            rows.append([family, generation, name, params, arch, release, variant,
                         bench, config, score, source])

# ============================= QWEN =============================

# --- Qwen (v1) Chat ---
m = [("Qwen-1.8B-Chat", 1.8, "Dense", "2023-11", "Instruct/chat"),
     ("Qwen-7B-Chat",   7,   "Dense", "2023-08", "Instruct/chat"),
     ("Qwen-14B-Chat",  14,  "Dense", "2023-09", "Instruct/chat")]
b = [("MMLU", "5-shot", [45.3, 58.2, 66.3]),
     ("C-Eval", "5-shot", [56.1, 63.5, 72.1]),
     ("GSM8K", "8-shot", [32.3, 51.7, 61.3]),
     ("MATH", "4-shot", [2.3, 11.6, 24.8]),
     ("HumanEval", "0-shot", [15.2, 29.9, 32.3]),
     ("MBPP", "3-shot", [14.2, 31.6, 40.8]),
     ("BBH", "3-shot", [22.3, 45.0, 53.4]),
     ("CMMLU", "5-shot", [52.1, 62.2, 71.0])]
add_gen(qwen_rows, "Qwen", "Qwen (v1)", m, b, "https://github.com/QwenLM/Qwen")

# --- Qwen1.5 Chat (only MT-Bench / AlpacaEval published for chat) ---
m = [("Qwen1.5-7B-Chat",  7,  "Dense", "2024-02", "Instruct/chat"),
     ("Qwen1.5-14B-Chat", 14, "Dense", "2024-02", "Instruct/chat"),
     ("Qwen1.5-32B-Chat", 32, "Dense", "2024-04", "Instruct/chat")]
b = [("MT-Bench", "avg score, GPT-4 judge", [7.60, 7.91, 8.30]),
     ("AlpacaEval 2.0", "length-controlled win rate", [13.20, 19.7, 27.49])]
add_gen(qwen_rows, "Qwen", "Qwen1.5", m, b,
        "https://qwenlm.github.io/blog/qwen1.5/ ; https://qwenlm.github.io/blog/qwen1.5-32b/")

# --- Qwen1.5 Base (fallback: no official chat academic scores exist) ---
m = [("Qwen1.5-1.8B", 1.8, "Dense", "2024-02", "Base (no official chat academic scores)"),
     ("Qwen1.5-4B",   4,   "Dense", "2024-02", "Base (no official chat academic scores)"),
     ("Qwen1.5-7B",   7,   "Dense", "2024-02", "Base (no official chat academic scores)"),
     ("Qwen1.5-14B",  14,  "Dense", "2024-02", "Base (no official chat academic scores)"),
     ("Qwen1.5-32B",  32,  "Dense", "2024-04", "Base (no official chat academic scores)")]
b = [("MMLU", "", [46.8, 56.1, 61.0, 67.6, 73.4]),
     ("C-Eval", "", [59.7, 67.6, 74.1, 78.7, 83.5]),
     ("GSM8K", "", [38.4, 57.0, 62.5, 70.1, 77.4]),
     ("MATH", "", [10.1, 10.0, 20.3, 29.2, 36.1]),
     ("HumanEval", "", [20.1, 25.6, 36.0, 37.8, 37.2]),
     ("MBPP", "", [18.0, 29.2, 37.4, 44.0, 49.4]),
     ("BBH", "", [24.2, 32.5, 40.2, 53.7, 66.8]),
     ("CMMLU", "", [57.8, 66.7, 73.1, 77.6, 82.3])]
add_gen(qwen_rows, "Qwen", "Qwen1.5", m, b,
        "https://qwenlm.github.io/blog/qwen1.5/ ; https://qwenlm.github.io/blog/qwen1.5-32b/")

# --- Qwen2-Instruct ---
m = [("Qwen2-0.5B-Instruct", 0.5, "Dense", "2024-06", "Instruct/chat"),
     ("Qwen2-1.5B-Instruct", 1.5, "Dense", "2024-06", "Instruct/chat"),
     ("Qwen2-7B-Instruct",   7,   "Dense", "2024-06", "Instruct/chat")]
b = [("MMLU", "", [37.9, 52.4, 70.5]),
     ("MMLU-Pro", "", [None, None, 44.1]),
     ("GPQA", "", [None, None, 25.3]),
     ("TheoremQA", "", [None, None, 25.3]),
     ("MT-Bench", "", [None, None, 8.41]),
     ("HumanEval", "0-shot", [17.1, 37.8, 79.9]),
     ("MBPP", "", [None, None, 67.2]),
     ("MultiPL-E", "", [None, None, 59.1]),
     ("EvalPlus", "", [None, None, 70.3]),
     ("LiveCodeBench", "", [None, None, 26.6]),
     ("GSM8K", "", [40.1, 61.6, 82.3]),
     ("MATH", "", [None, None, 49.6]),
     ("C-Eval", "", [45.2, 63.8, 77.2]),
     ("IFEval", "prompt strict-acc.", [20.0, 29.0, None]),
     ("AlignBench", "", [None, None, 7.21])]
add_gen(qwen_rows, "Qwen", "Qwen2", m, b, "https://qwenlm.github.io/blog/qwen2/")

# --- Qwen2.5-Instruct ---
m = [("Qwen2.5-0.5B-Instruct", 0.5, "Dense", "2024-09", "Instruct/chat"),
     ("Qwen2.5-1.5B-Instruct", 1.5, "Dense", "2024-09", "Instruct/chat"),
     ("Qwen2.5-3B-Instruct",   3,   "Dense", "2024-09", "Instruct/chat"),
     ("Qwen2.5-7B-Instruct",   7,   "Dense", "2024-09", "Instruct/chat"),
     ("Qwen2.5-14B-Instruct",  14,  "Dense", "2024-09", "Instruct/chat"),
     ("Qwen2.5-32B-Instruct",  32,  "Dense", "2024-09", "Instruct/chat")]
b = [("MMLU-Pro", "", [15.0, 32.4, 43.7, 56.3, 63.7, 69.0]),
     ("MMLU-Redux", "", [24.1, 50.7, 64.4, 75.4, 80.0, 83.9]),
     ("GPQA", "", [29.8, 29.8, 30.3, 36.4, 45.5, 49.5]),
     ("MATH", "", [34.4, 55.2, 65.9, 75.5, 80.0, 83.1]),
     ("GSM8K", "", [49.6, 73.2, 86.7, 91.6, 94.8, 95.9]),
     ("HumanEval", "", [35.4, 61.6, 74.4, 84.8, 83.5, 88.4]),
     ("MBPP", "", [49.6, 63.2, 72.7, 79.2, 82.0, 84.0]),
     ("MultiPL-E", "", [28.5, 50.4, 60.2, 70.4, 72.8, 75.4]),
     ("LiveCodeBench", "questions 2305-2409", [5.1, 14.8, 19.9, 28.7, 42.6, 51.2]),
     ("LiveBench", "release 0831", [12.6, 18.8, 26.8, 35.9, 44.4, 50.7]),
     ("IFEval", "strict-prompt", [27.9, 42.5, 58.2, 71.2, 81.0, 79.5]),
     ("Arena-Hard", "", [None, None, None, 52.0, 68.3, 74.5]),
     ("AlignBench v1.1", "", [None, None, None, 7.33, 7.94, 7.93]),
     ("MT-Bench", "", [None, None, None, 8.75, 8.88, 9.20])]
add_gen(qwen_rows, "Qwen", "Qwen2.5", m, b, "https://qwenlm.github.io/blog/qwen2.5-llm/")

# --- Qwen3 (thinking / non-thinking), tech report arXiv 2505.09388 ---
q3_src = "Qwen3 Technical Report, arXiv:2505.09388 (blog tables are images)"
q3_models = [("Qwen3-0.6B", 0.6, "Dense", "2025-04"),
             ("Qwen3-1.7B", 1.7, "Dense", "2025-04"),
             ("Qwen3-4B",   4,   "Dense", "2025-04"),
             ("Qwen3-8B",   8,   "Dense", "2025-04"),
             ("Qwen3-14B",  14,  "Dense", "2025-04"),
             ("Qwen3-32B",  32,  "Dense", "2025-04"),
             ("Qwen3-30B-A3B", 30, "MoE (30B total / 3B active)", "2025-04")]
q3_thinking = [
    ("MMLU-Redux", "", [55.6, 73.9, 83.7, 87.5, 88.6, 90.9, 89.5]),
    ("GPQA-Diamond", "", [27.9, 40.1, 55.9, 62.0, 64.0, 68.4, 65.8]),
    ("C-Eval", "", [50.4, 68.1, 77.5, 83.4, 86.2, 87.3, 86.6]),
    ("AIME'24", "", [10.7, 48.3, 73.8, 76.0, 79.3, 81.4, 80.4]),
    ("AIME'25", "", [15.1, 36.8, 65.6, 67.3, 70.4, 72.9, 70.9]),
    ("MATH-500", "", [77.6, 93.4, 97.0, 97.4, 96.8, 97.2, 98.0]),
    ("LiveCodeBench", "v5", [12.3, 33.2, 54.2, 57.5, 63.5, 65.7, 62.6]),
    ("CodeForces rating", "", [None, None, 1671, 1785, 1766, 1977, 1974]),
    ("CodeForces percentile", "", [None, None, 92.8, 95.6, 95.3, 97.7, 97.7]),
    ("IFEval", "", [59.2, 72.5, 81.9, 85.0, 85.4, 85.0, 86.5]),
    ("Arena-Hard", "", [8.5, 43.1, 76.6, 85.8, 91.7, 93.8, 91.0]),
    ("LiveBench", "", [30.3, 51.1, 63.6, 67.1, 71.3, 74.9, 74.3]),
    ("BFCL v3", "", [46.4, 56.6, 65.9, 68.1, 70.4, 70.3, 69.1]),
    ("Multi-IF", "", [36.1, 51.2, 66.3, 71.2, 74.8, 73.0, 72.2])]
q3_nonthinking = [
    ("MMLU-Redux", "", [44.6, 64.4, 77.3, 79.5, 82.0, 85.7, 84.1]),
    ("GPQA-Diamond", "", [22.9, 28.6, 41.7, 39.3, 54.8, 54.6, 54.8]),
    ("C-Eval", "", [42.6, 61.0, 72.2, 77.9, 81.0, 83.3, 82.9]),
    ("AIME'24", "", [3.4, 13.4, 25.0, 29.1, 31.7, 31.0, 32.8]),
    ("AIME'25", "", [2.6, 9.8, 19.1, 20.9, 23.3, 20.2, 21.6]),
    ("MATH-500", "", [55.2, 73.0, 84.8, 87.4, 90.0, 88.6, 89.8]),
    ("LiveCodeBench", "v5", [3.6, 11.6, 21.3, 22.8, 29.0, 31.3, 29.8]),
    ("CodeForces rating", "", [None, None, 842, 1110, 1200, 1353, 1267]),
    ("CodeForces percentile", "", [None, None, 33.7, 52.4, 58.6, 71.0, 64.1]),
    ("IFEval", "", [54.5, 68.2, 81.2, 83.0, 84.8, 83.2, 83.7]),
    ("Arena-Hard", "", [6.5, 36.9, 66.2, 79.6, 86.3, 92.8, 88.0]),
    ("LiveBench", "", [21.8, 35.6, 48.4, 53.5, 59.6, 59.8, 59.4]),
    ("BFCL v3", "", [44.1, 52.2, 57.6, 60.2, 61.5, 63.0, 58.6]),
    ("Multi-IF", "", [33.3, 44.7, 61.3, 69.2, 72.9, 70.7, 70.8])]
add_gen(qwen_rows, "Qwen", "Qwen3",
        [(n, p, a, r, "Thinking mode") for n, p, a, r in q3_models], q3_thinking, q3_src)
add_gen(qwen_rows, "Qwen", "Qwen3",
        [(n, p, a, r, "Non-thinking mode") for n, p, a, r in q3_models], q3_nonthinking, q3_src)

# --- Qwen3-2507 updates ---
m = [("Qwen3-4B-Instruct-2507",       4,  "Dense", "2025-08", "Instruct (non-thinking)"),
     ("Qwen3-4B-Thinking-2507",       4,  "Dense", "2025-08", "Thinking"),
     ("Qwen3-30B-A3B-Instruct-2507",  30, "MoE (30B total / 3B active)", "2025-07", "Instruct (non-thinking)"),
     ("Qwen3-30B-A3B-Thinking-2507",  30, "MoE (30B total / 3B active)", "2025-07", "Thinking")]
b = [("MMLU-Pro", "", [69.6, 74.0, 78.4, 80.9]),
     ("MMLU-Redux", "", [84.2, 86.1, 89.3, 91.4]),
     ("GPQA", "", [62.0, 65.8, 70.4, 73.4]),
     ("SuperGPQA", "", [42.8, 47.8, 53.4, 56.8]),
     ("AIME'25", "", [47.4, 81.3, 61.3, 85.0]),
     ("HMMT'25", "", [31.0, 55.5, 43.0, 71.4]),
     ("ZebraLogic", "", [80.2, None, 90.0, None]),
     ("LiveBench", "release 20241125", [63.0, 71.8, 69.0, 76.8]),
     ("LiveCodeBench", "v6, questions 25.02-25.05", [35.1, 55.2, 43.2, 66.0]),
     ("CFEval", "", [None, 1852, None, 2044]),
     ("OJBench", "", [None, 17.9, None, 25.1]),
     ("MultiPL-E", "", [76.8, None, 83.8, None]),
     ("Aider-Polyglot", "", [12.9, None, 35.6, None]),
     ("IFEval", "", [83.4, 87.4, 84.7, 88.9]),
     ("Arena-Hard v2", "win rate, GPT-4.1 judge", [43.4, 34.9, 69.0, 56.0]),
     ("Creative Writing v3", "", [83.5, 75.6, 86.0, 84.4]),
     ("WritingBench", "", [83.4, 83.3, 85.5, 85.0]),
     ("BFCL v3", "", [61.9, 71.2, 65.1, 72.4]),
     ("TAU1-Retail", "", [48.7, 66.1, 59.1, 67.8]),
     ("TAU1-Airline", "", [32.0, 48.0, 40.0, 48.0]),
     ("TAU2-Retail", "", [40.4, 53.5, 57.0, 58.8]),
     ("TAU2-Airline", "", [24.0, 58.0, 38.0, 58.0]),
     ("TAU2-Telecom", "", [13.2, 27.2, 12.3, 26.3]),
     ("Multi-IF", "", [69.0, 77.3, 67.9, 76.4]),
     ("MMLU-ProX", "", [61.6, 64.2, 72.0, 76.4]),
     ("INCLUDE", "", [60.1, 64.4, 71.9, 74.4]),
     ("PolyMATH", "", [31.1, 46.2, 43.1, 52.6])]
add_gen(qwen_rows, "Qwen", "Qwen3-2507", m, b,
        "HF model cards: huggingface.co/Qwen/Qwen3-{4B,30B-A3B}-{Instruct,Thinking}-2507")

# --- Qwen3.5 (from live HF model cards, extracted 2026-08) ---
q35_note = "Live HF model card (extracted 2026-08)"
m = [("Qwen3.5-4B", 4, "Dense", "2026-02", "Thinking (default)"),
     ("Qwen3.5-9B", 9, "Dense", "2026-02", "Thinking (default)")]
b = [("MMLU-Pro", "", [79.1, 82.5]),
     ("MMLU-Redux", "", [88.8, 91.1]),
     ("C-Eval", "", [85.1, 88.2]),
     ("SuperGPQA", "", [52.9, 58.2]),
     ("GPQA-Diamond", "", [76.2, 81.7]),
     ("IFEval", "", [89.8, 91.5]),
     ("IFBench", "", [59.2, 64.5]),
     ("MultiChallenge", "", [49.0, 54.5]),
     ("AA-LCR", "", [57.0, 63.0]),
     ("LongBench v2", "", [50.0, 55.2]),
     ("HMMT Feb'25", "", [74.0, 83.2]),
     ("HMMT Nov'25", "", [76.8, 82.9]),
     ("LiveCodeBench", "v6", [55.8, 65.6]),
     ("OJBench", "", [24.1, 29.2]),
     ("BFCL v4", "", [50.3, 66.1]),
     ("TAU2-Bench", "", [79.9, 79.1]),
     ("MMLU-ProX", "", [71.5, 76.3]),
     ("INCLUDE", "", [71.0, 75.6]),
     ("PolyMATH", "", [51.1, 57.3])]
add_gen(qwen_rows, "Qwen", "Qwen3.5", m, b,
        "huggingface.co/Qwen/Qwen3.5-4B ; /Qwen3.5-9B (" + q35_note + ")")

m = [("Qwen3.5-27B", 27, "Dense (multimodal)", "2026-02", "Thinking (default)")]
b = [("MMLU-Pro", "", [86.1]), ("MMLU-Redux", "", [93.2]), ("C-Eval", "", [90.5]),
     ("SuperGPQA", "", [65.6]), ("GPQA-Diamond", "", [85.5]), ("IFEval", "", [95.0]),
     ("LiveCodeBench", "v6", [80.7]), ("SWE-bench Verified", "", [72.4]),
     ("HMMT Feb'25", "", [92.0]), ("AIME'26", "", [92.6]),
     ("MMMU-Pro", "vision", [75.0]), ("MathVision", "vision", [86.0]),
     ("MMStar", "vision", [81.0]), ("VideoMME", "w/ subtitles", [87.0]),
     ("ScreenSpot Pro", "vision", [70.3])]
add_gen(qwen_rows, "Qwen", "Qwen3.5", m, b,
        "huggingface.co/Qwen/Qwen3.5-27B (" + q35_note + ")")

m = [("Qwen3.5-35B-A3B", 35, "MoE (35B total / 3B active, multimodal)", "2026-02", "Thinking (default)")]
b = [("MMLU-Pro", "", [85.3]), ("MMLU-Redux", "", [93.3]), ("C-Eval", "", [90.2]),
     ("SuperGPQA", "", [63.4]), ("GPQA-Diamond", "", [84.2]),
     ("LiveCodeBench", "v6", [74.6]), ("SWE-bench Verified", "", [70.0]),
     ("HMMT Feb'25", "", [89.0]), ("AIME'26", "", [91.0]),
     ("MMMU", "vision", [81.4]), ("MMMU-Pro", "vision", [75.1])]
add_gen(qwen_rows, "Qwen", "Qwen3.5", m, b,
        "huggingface.co/Qwen/Qwen3.5-35B-A3B (" + q35_note + ")")

m = [("Qwen3.5-0.8B", 0.8, "Dense", "2026-02", "Non-thinking rows"),
     ("Qwen3.5-2B",   2,   "Dense", "2026-02", "Non-thinking rows")]
b = [("MMLU-Pro", "", [29.7, 55.3]),
     ("MMLU-Redux", "", [48.5, 69.2]),
     ("C-Eval", "", [46.4, 65.2]),
     ("SuperGPQA", "", [16.9, 30.4]),
     ("IFEval", "", [52.1, 61.2]),
     ("MMMLU", "", [34.1, 56.9])]
add_gen(qwen_rows, "Qwen", "Qwen3.5", m, b,
        "huggingface.co/Qwen/Qwen3.5-0.8B (" + q35_note + "; thinking-mode rows exist on card, not transcribed)")

# --- Qwen3.6 (from live HF model cards, extracted 2026-08) ---
m = [("Qwen3.6-27B", 27, "Dense (multimodal)", "2026-04", "Thinking (default)")]
b = [("MMLU-Pro", "", [86.2]), ("MMLU-Redux", "", [93.5]), ("SuperGPQA", "", [66.0]),
     ("C-Eval", "", [91.4]), ("GPQA-Diamond", "", [87.8]), ("HLE", "no tools", [24.0]),
     ("LiveCodeBench", "v6", [83.9]), ("HMMT Feb'25", "", [93.8]),
     ("HMMT Nov'25", "", [90.7]), ("HMMT Feb'26", "", [84.3]),
     ("IMOAnswerBench", "", [80.8]), ("AIME'26", "", [94.1]),
     ("SWE-bench Verified", "", [77.2]), ("SWE-bench Pro", "", [53.5]),
     ("Terminal-Bench 2.0", "", [59.3])]
add_gen(qwen_rows, "Qwen", "Qwen3.6", m, b,
        "huggingface.co/Qwen/Qwen3.6-27B (live card, extracted 2026-08)")

m = [("Qwen3.6-35B-A3B", 35, "MoE (35B total / 3B active, multimodal)", "2026-04", "Thinking (default)")]
b = [("MMLU-Pro", "", [85.2]), ("MMLU-Redux", "", [93.3]), ("SuperGPQA", "", [64.7]),
     ("C-Eval", "", [90.0]), ("GPQA-Diamond", "", [86.0]), ("HLE", "no tools", [21.4]),
     ("LiveCodeBench", "v6", [80.4]), ("HMMT Feb'25", "", [90.7]),
     ("HMMT Nov'25", "", [89.1]), ("HMMT Feb'26", "", [83.6]),
     ("IMOAnswerBench", "", [78.9]), ("AIME'26", "", [92.7]),
     ("SWE-bench Verified", "", [73.4]), ("Terminal-Bench 2.0", "", [51.5])]
add_gen(qwen_rows, "Qwen", "Qwen3.6", m, b,
        "huggingface.co/Qwen/Qwen3.6-35B-A3B (live card, extracted 2026-08)")

# ============================= GEMMA =============================

# --- Gemma 1 PT ---
g1_src = "ai.google.dev/gemma/docs/core/model_card ; arXiv:2403.08295"
m = [("Gemma 2B",  2, "Dense", "2024-02", "Pretrained (PT)"),
     ("Gemma 7B",  7, "Dense", "2024-02", "Pretrained (PT)")]
b = [("MMLU", "5-shot, top-1", [42.3, 64.3]),
     ("HellaSwag", "0-shot", [71.4, 81.2]),
     ("PIQA", "0-shot", [77.3, 81.2]),
     ("SocialIQA", "0-shot", [49.7, 51.8]),
     ("BoolQ", "0-shot", [69.4, 83.2]),
     ("WinoGrande", "partial score", [65.4, 72.3]),
     ("CommonsenseQA", "7-shot", [65.3, 71.3]),
     ("OpenBookQA", "", [47.8, 52.8]),
     ("ARC-e", "", [73.2, 81.5]),
     ("ARC-c", "", [42.1, 53.2]),
     ("TriviaQA", "5-shot", [53.2, 63.4]),
     ("Natural Questions", "5-shot", [12.5, 23.0]),
     ("HumanEval", "pass@1", [22.0, 32.3]),
     ("MBPP", "3-shot", [29.2, 44.4]),
     ("GSM8K", "maj@1", [17.7, 46.4]),
     ("MATH", "4-shot", [11.8, 24.3]),
     ("AGIEval", "", [24.2, 41.7]),
     ("BIG-Bench", "", [35.2, 55.1]),
     ("Average (official table)", "", [44.9, 56.4])]
add_gen(gemma_rows, "Gemma", "Gemma 1", m, b, g1_src)

# Gemma 1 IT: no academic table published; official human-pref win rates (tech report v4, reflects 1.1)
m = [("Gemma 2B IT", 2, "Dense", "2024-02", "Instruction-tuned (IT)"),
     ("Gemma 7B IT", 7, "Dense", "2024-02", "Instruction-tuned (IT)")]
b = [("Human-pref win rate vs Mistral-7B-Instruct-v0.2", "instruction following", [45.0, 61.2]),
     ("Human-pref win rate vs Mistral-7B-Instruct-v0.2", "safety", [60.1, 63.5])]
add_gen(gemma_rows, "Gemma", "Gemma 1", m, b,
        g1_src + " (no academic IT benchmark table was ever published for Gemma 1 / 1.1 IT)")

# --- Gemma 2 PT ---
g2_src = "ai.google.dev/gemma/docs/core/model_card_2 ; arXiv:2408.00118"
m = [("Gemma 2 2B",  2,  "Dense", "2024-07", "Pretrained (PT)"),
     ("Gemma 2 9B",  9,  "Dense", "2024-06", "Pretrained (PT)"),
     ("Gemma 2 27B", 27, "Dense", "2024-06", "Pretrained (PT)")]
b = [("MMLU", "5-shot, top-1", [51.3, 71.3, 75.2]),
     ("HellaSwag", "10-shot", [73.0, 81.9, 86.4]),
     ("PIQA", "0-shot", [77.8, 81.7, 83.2]),
     ("SocialIQA", "0-shot", [51.9, 53.4, 53.7]),
     ("BoolQ", "0-shot", [72.5, 84.2, 84.8]),
     ("WinoGrande", "partial score", [70.9, 80.6, 83.7]),
     ("ARC-e", "0-shot", [80.1, 88.0, 88.6]),
     ("ARC-c", "25-shot", [55.4, 68.4, 71.4]),
     ("TriviaQA", "5-shot", [59.4, 76.6, 83.7]),
     ("Natural Questions", "5-shot", [16.7, 29.2, 34.5]),
     ("HumanEval", "pass@1", [17.7, 40.2, 51.8]),
     ("MBPP", "3-shot", [29.6, 52.4, 62.6]),
     ("GSM8K", "5-shot, maj@1", [23.9, 68.6, 74.0]),
     ("MATH", "4-shot", [15.0, 36.6, 42.3]),
     ("AGIEval", "3-5-shot", [30.6, 52.8, 55.1]),
     ("DROP", "3-shot, F1", [52.0, 69.4, 72.2]),
     ("BIG-Bench", "3-shot, CoT", [41.9, 68.2, 74.9])]
add_gen(gemma_rows, "Gemma", "Gemma 2", m, b,
        g2_src + " (tech report Table 13 has slightly different 2B values: MMLU 52.2, GSM8K 24.3, MBPP 30.2)")

# Gemma 2 IT: only Arena Elo published
m = [("Gemma 2 2B IT",  2,  "Dense", "2024-07", "Instruction-tuned (IT)"),
     ("Gemma 2 9B IT",  9,  "Dense", "2024-06", "Instruction-tuned (IT)"),
     ("Gemma 2 27B IT", 27, "Dense", "2024-06", "Instruction-tuned (IT)")]
b = [("LMSYS Chatbot Arena Elo", "as cited in tech report", [1126, 1187, 1218])]
add_gen(gemma_rows, "Gemma", "Gemma 2", m, b,
        g2_src + " (no academic IT benchmark table was ever published for Gemma 2 IT)")

# --- Gemma 3 PT ---
g3_src = "ai.google.dev/gemma/docs/core/model_card_3 ; arXiv:2503.19786"
m = [("Gemma 3 1B",  1,  "Dense", "2025-03", "Pretrained (PT)"),
     ("Gemma 3 4B",  4,  "Dense (multimodal)", "2025-03", "Pretrained (PT)"),
     ("Gemma 3 12B", 12, "Dense (multimodal)", "2025-03", "Pretrained (PT)"),
     ("Gemma 3 27B", 27, "Dense (multimodal)", "2025-03", "Pretrained (PT)")]
b = [("HellaSwag", "10-shot", [62.3, 77.2, 84.2, 85.6]),
     ("BoolQ", "0-shot", [63.2, 72.3, 78.8, 82.4]),
     ("PIQA", "0-shot", [73.8, 79.6, 81.8, 83.3]),
     ("SocialIQA", "0-shot", [48.9, 51.9, 53.4, 54.9]),
     ("TriviaQA", "5-shot", [39.8, 65.8, 78.2, 85.5]),
     ("Natural Questions", "5-shot", [9.48, 20.0, 31.4, 36.1]),
     ("ARC-c", "25-shot", [38.4, 56.2, 68.9, 70.6]),
     ("ARC-e", "0-shot", [73.0, 82.4, 88.3, 89.0]),
     ("WinoGrande", "5-shot", [58.2, 64.7, 74.3, 78.8]),
     ("BIG-Bench Hard", "few-shot", [28.4, 50.9, 72.6, 77.7]),
     ("DROP", "1-shot, F1", [42.4, 60.1, 72.2, 77.2]),
     ("MMLU", "5-shot", [None, 59.6, 74.5, 78.6]),
     ("MMLU-Pro", "5-shot, CoT", [None, 29.2, 45.3, 52.2]),
     ("AGIEval", "3-5-shot", [None, 42.1, 57.4, 66.2]),
     ("MATH", "4-shot", [None, 24.2, 43.3, 50.0]),
     ("GSM8K", "8-shot", [None, 38.4, 71.0, 82.6]),
     ("GPQA", "5-shot", [None, 15.0, 25.4, 24.3]),
     ("MBPP", "3-shot", [None, 46.0, 60.4, 65.6]),
     ("HumanEval", "0-shot", [None, 36.0, 45.7, 48.8]),
     ("MGSM", "", [2.04, 34.7, 64.3, 74.3]),
     ("Global-MMLU-Lite", "", [24.9, 57.0, 69.4, 75.7]),
     ("WMT24++", "ChrF", [36.7, 48.4, 53.9, 55.7]),
     ("FloRes", "", [29.5, 39.2, 46.0, 48.8]),
     ("XQuAD", "", [43.9, 68.0, 74.5, 76.8]),
     ("ECLeKTic", "", [4.69, 11.0, 17.2, 24.4]),
     ("IndicGenBench", "", [41.4, 57.2, 61.7, 63.4])]
add_gen(gemma_rows, "Gemma", "Gemma 3", m, b, g3_src)

# --- Gemma 3 IT ---
m = [("Gemma 3 1B IT",  1,  "Dense", "2025-03", "Instruction-tuned (IT)"),
     ("Gemma 3 4B IT",  4,  "Dense (multimodal)", "2025-03", "Instruction-tuned (IT)"),
     ("Gemma 3 12B IT", 12, "Dense (multimodal)", "2025-03", "Instruction-tuned (IT)"),
     ("Gemma 3 27B IT", 27, "Dense (multimodal)", "2025-03", "Instruction-tuned (IT)")]
b = [("MMLU-Pro", "0-shot", [14.7, 43.6, 60.6, 67.5]),
     ("LiveCodeBench", "0-shot", [1.9, 12.6, 24.6, 29.7]),
     ("Bird-SQL (dev)", "", [6.4, 36.3, 47.9, 54.4]),
     ("GPQA-Diamond", "0-shot", [19.2, 30.8, 40.9, 42.4]),
     ("SimpleQA", "0-shot", [2.2, 4.0, 6.3, 10.0]),
     ("FACTS Grounding", "", [36.4, 70.1, 75.8, 74.9]),
     ("Global-MMLU-Lite", "", [34.2, 54.5, 69.5, 75.1]),
     ("MATH", "0-shot", [48.0, 75.6, 83.8, 89.0]),
     ("HiddenMath", "0-shot", [15.8, 43.0, 54.5, 60.3]),
     ("MMMU (val)", "vision", [None, 48.8, 59.6, 64.9]),
     ("BIG-Bench Hard", "0-shot", [39.1, 72.2, 85.7, 87.6]),
     ("HumanEval", "0-shot", [41.5, 71.3, 85.4, 87.8]),
     ("GSM8K", "0-shot", [62.8, 89.2, 94.4, 95.9]),
     ("DocVQA", "vision", [None, 75.8, 87.1, 86.6]),
     ("VQAv2", "vision", [None, 62.4, 71.6, 71.0]),
     ("ChartQA", "vision", [None, 68.8, 75.7, 78.0]),
     ("LMArena Elo", "as cited in tech report (Mar 2025)", [None, None, None, 1338])]
add_gen(gemma_rows, "Gemma", "Gemma 3", m, b, g3_src)

# --- Gemma 3n ---
g3n_src = "ai.google.dev/gemma/docs/gemma-3n/model_card"
m = [("Gemma 3n E2B", 2, "Selective activation (~2B effective)", "2025-06", "Pretrained (PT)"),
     ("Gemma 3n E4B", 4, "Selective activation (~4B effective)", "2025-06", "Pretrained (PT)")]
b = [("HellaSwag", "10-shot", [72.2, 78.6]),
     ("BoolQ", "0-shot", [76.4, 81.6]),
     ("PIQA", "0-shot", [78.9, 81.0]),
     ("SocialIQA", "0-shot", [48.8, 50.0]),
     ("TriviaQA", "5-shot", [60.8, 70.2]),
     ("Natural Questions", "5-shot", [15.5, 20.9]),
     ("ARC-c", "25-shot", [51.7, 61.6]),
     ("ARC-e", "0-shot", [75.8, 81.6]),
     ("WinoGrande", "5-shot", [66.8, 71.7]),
     ("BIG-Bench Hard", "few-shot", [44.3, 52.9]),
     ("DROP", "1-shot, F1", [53.9, 60.8])]
add_gen(gemma_rows, "Gemma", "Gemma 3n", m, b, g3n_src)

m = [("Gemma 3n E2B IT", 2, "Selective activation (~2B effective)", "2025-06", "Instruction-tuned (IT)"),
     ("Gemma 3n E4B IT", 4, "Selective activation (~4B effective)", "2025-06", "Instruction-tuned (IT)")]
b = [("MMLU", "0-shot", [60.1, 64.9]),
     ("MMLU-Pro", "0-shot", [40.5, 50.6]),
     ("GPQA-Diamond", "relaxed accuracy", [24.8, 23.7]),
     ("HumanEval", "pass@1", [66.5, 75.0]),
     ("MBPP", "3-shot", [56.6, 63.6]),
     ("LiveCodeBench", "v5", [18.6, 25.7]),
     ("AIME'25", "", [6.7, 11.6]),
     ("HiddenMath", "", [27.7, 37.7]),
     ("Global-MMLU-Lite", "", [59.0, 64.5]),
     ("MGSM", "", [53.1, 60.7]),
     ("WMT24++", "ChrF", [42.7, 50.1]),
     ("ECLeKTic", "", [2.5, 1.9])]
add_gen(gemma_rows, "Gemma", "Gemma 3n", m, b, g3n_src)

# --- Gemma 4 (IT "Thinking" only; no PT table published; live cards extracted 2026-08) ---
g4_src = ("ai.google.dev/gemma/docs/core/model_card_4 ; huggingface.co/google/gemma-4-E4B ; "
          "deepmind.google/models/gemma/gemma-4/ ; arXiv:2607.02770 (live pages, extracted 2026-08)")
m = [("Gemma 4 E2B",     2.3,  "Selective activation (2.3B effective / 5.1B incl. embeddings)", "2026-03", "IT, Thinking mode"),
     ("Gemma 4 E4B",     4.5,  "Selective activation (4.5B effective / 8B incl. embeddings)",   "2026-03", "IT, Thinking mode"),
     ("Gemma 4 12B",     11.95,"Dense (unified)", "2026-06", "IT, Thinking mode"),
     ("Gemma 4 26B A4B", 25.2, "MoE (25.2B total / 3.8B active)", "2026-03", "IT, Thinking mode"),
     ("Gemma 4 31B",     30.7, "Dense", "2026-03", "IT, Thinking mode")]
b = [("MMLU-Pro", "", [60.0, 69.4, 77.2, 82.6, 85.2]),
     ("MMMLU (multilingual MMLU)", "", [67.4, 76.6, 83.4, 86.3, 88.4]),
     ("GPQA-Diamond", "no tools", [43.4, 58.6, 78.8, 82.3, 84.3]),
     ("AIME'26", "no tools", [37.5, 42.5, 77.5, 88.3, 89.2]),
     ("LiveCodeBench", "v6", [44.0, 52.0, 72.0, 77.1, 80.0]),
     ("CodeForces Elo", "", [633, 940, 1659, 1718, 2150]),
     ("BigBench Extra Hard", "", [21.9, 33.1, 53.0, 64.8, 74.4]),
     ("Tau2-bench", "avg of 3 domains", [24.5, 42.2, 69.0, 68.2, 76.9]),
     ("Tau2-bench Retail", "DeepMind page", [29.4, 57.5, None, 85.5, 86.4]),
     ("HLE", "no tools", [None, None, 5.2, 8.7, 19.5]),
     ("HLE", "with search", [None, None, None, 17.2, 26.5]),
     ("MMMU-Pro", "vision", [44.2, 52.6, 69.1, 73.8, 76.9]),
     ("MATH-Vision", "vision", [52.4, 59.5, 79.7, 82.4, 85.6]),
     ("MedXPertQA MM", "vision", [23.5, 28.7, 48.7, 58.1, 61.3]),
     ("OmniDocBench 1.5", "lower is better", [0.290, 0.181, 0.164, 0.149, 0.131]),
     ("CoVoST", "audio; 12B excl. zh", [33.47, 35.54, 38.5, None, None]),
     ("FLEURS", "audio, lower is better; 12B excl. zh", [0.09, 0.08, 0.069, None, None]),
     ("MRCR v2", "8-needle, 128k context", [19.1, 25.4, 43.4, 44.1, 66.4]),
     ("LMArena Elo", "as of 2026-04-02", [None, None, None, 1441, 1452])]
add_gen(gemma_rows, "Gemma", "Gemma 4", m, b, g4_src)

# ============================= WORKBOOK =============================

wb = Workbook()
FONT = "Arial"
header_font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
header_fill = PatternFill("solid", fgColor="1F3864")
body_font = Font(name=FONT, size=10)
bold_font = Font(name=FONT, size=10, bold=True)
title_font = Font(name=FONT, size=14, bold=True, color="1F3864")
h2_font = Font(name=FONT, size=11, bold=True, color="1F3864")
wrap = Alignment(wrap_text=True, vertical="top")
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_table(ws, n_rows, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2, max_row=n_rows + 1, max_col=len(widths)):
        for c in row:
            c.font = body_font
            c.border = border
            if c.column_letter in ("I", "K"):
                c.alignment = wrap
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(widths))}{n_rows + 1}"

# README sheet
ws = wb.active
ws.title = "README"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 118
readme = [
    ("Qwen & Gemma official benchmark scores — models up to 35B parameters, Aug 2023 – Aug 2026", "title"),
    ("", None),
    ("Compiled 2026-08-09 from official sources only: vendor blogs, technical reports (arXiv), GitHub READMEs, and Hugging Face model cards. Every score is transcribed as published; nothing is interpolated or re-run.", None),
    ("", None),
    ("Sheets", "h2"),
    ("Qwen scores / Gemma scores — tidy long format (one row per model x benchmark). Filter or pivot on any column. 'Source' gives the exact origin of each generation's numbers.", None),
    ("Third-party sources — a survey of independent leaderboards covering these models, with coverage notes and download routes, per your request to scope that before collecting.", None),
    ("", None),
    ("Key caveats for cross-generation comparison", "h2"),
    ("1. Benchmark suites changed almost every generation. Qwen dropped classic MMLU, GSM8K, HumanEval, MBPP and MT-Bench after Qwen2.5, moving to MMLU-Redux/Pro, AIME, LiveCodeBench, Arena-Hard. Gemma's IT suite changed completely between Gemma 3 and Gemma 4. Very few benchmarks form a continuous series across all generations.", None),
    ("2. Most continuous series to track: Qwen — GPQA(-Diamond) (Qwen2.5 on), MMLU-Pro / MMLU-Redux (Qwen2.5 on), IFEval (Qwen2 on), LiveCodeBench (version windows differ!), C-Eval (all generations). Gemma — MMLU-Pro and GPQA-Diamond IT scores (Gemma 3 / 3n / 4), plus the PT tables (Gemma 1-3 share HellaSwag, PIQA, BoolQ, TriviaQA, ARC, WinoGrande, GSM8K, MATH, HumanEval, MBPP).", None),
    ("3. Scores are NOT comparable across labs. Qwen and Google use different harnesses, prompts, and shot counts (e.g., Gemma 3 IT numbers are 0-shot; many Qwen numbers use few-shot or CoT setups).", None),
    ("4. LiveCodeBench rows use different question windows (2305-2409, v5, v6 25.02-25.05) — treat each version as a separate benchmark.", None),
    ("5. Thinking vs non-thinking: from Qwen3 and Gemma 4 onward, scores depend heavily on reasoning mode. The 'Variant / mode' column tracks this — compare like with like.", None),
    ("6. Gaps that cannot be filled (never published): Qwen1.5 chat academic scores (only MT-Bench/AlpacaEval; base-model rows included here as a labeled fallback); Gemma 1 & Gemma 2 IT academic tables (only Arena Elo / human-preference win rates); Gemma 4 pretrained scores; QwQ-32B scores (published only as chart images); Qwen2 0.5B/1.5B beyond the 5 benchmarks shown.", None),
    ("7. Qwen3.5, Qwen3.6, and Gemma 4 rows were extracted from live model cards in Aug 2026 and are flagged as such in the Source column; Qwen3.5-2B/0.8B thinking-mode rows and some agentic/vision rows were not fully transcribed.", None),
    ("8. MoE models (Qwen3-30B-A3B, Qwen3.5/3.6-35B-A3B, Gemma 4 26B A4B) are listed by total parameters with active counts in the Architecture column — decide whether to bucket by total or active params for size-scaling analysis.", None),
    ("9. Score units vary: most are accuracy/pass@1 percentages, but MT-Bench and AlignBench are 0-10 scales, Elo/CodeForces are ratings, AlpacaEval/Arena-Hard are win rates, OmniDocBench and FLEURS are lower-is-better. Check 'Eval config / notes'.", None),
]
r = 1
for text, kind in readme:
    c = ws.cell(row=r, column=1, value=text)
    c.alignment = wrap
    c.font = title_font if kind == "title" else h2_font if kind == "h2" else body_font
    r += 1

# Data sheets
widths = [8, 12, 28, 10, 34, 9, 26, 30, 26, 9, 60]
ws_q = wb.create_sheet("Qwen scores")
ws_q.append(HEADERS)
for row in qwen_rows:
    ws_q.append(row)
style_table(ws_q, len(qwen_rows), widths)

ws_g = wb.create_sheet("Gemma scores")
ws_g.append(HEADERS)
for row in gemma_rows:
    ws_g.append(row)
style_table(ws_g, len(gemma_rows), widths)

print(f"Qwen rows: {len(qwen_rows)}, Gemma rows: {len(gemma_rows)}")
wb.save("./qwen_gemma_benchmarks.xlsx")
print("saved")
