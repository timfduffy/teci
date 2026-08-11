#!/usr/bin/env python3
"""Add third-party sources survey sheet to the workbook."""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = load_workbook("./qwen_gemma_benchmarks.xlsx")
ws = wb.create_sheet("Third-party sources")

headers = ["Rank / verdict", "Source", "URL(s)", "What it measures",
           "Coverage of Qwen/Gemma ≤35B", "Data access", "Caveat for tracking progress over time"]
rows = [
["1 — best single source", "LMArena (Chatbot Arena)", "lmarena.ai/leaderboard/text (now arena.ai); archives: HF space lmarena-ai/chatbot-arena-leaderboard (elo_results_*.pkl, 2023-2024); HF dataset mathewhe/chatbot-arena-elo (daily snapshots); github.com/oolong-tea-2026/arena-ai-leaderboards (daily JSON + API)",
 "Human-preference Elo from pairwise battles",
 "Only source spanning all generations 2023-2026 on one scale; deprecated models retained. Spot-checked: gemma-7b-it 1137, gemma-2-2b/9b/27b-it 1200/1267/1289, gemma-3-4b/12b/27b-it 1303/1342/1366, qwen1.5-7b-chat 1143, qwen1.5-32b-chat 1203, qwen3-32b 1347, qwen3-30b-a3b-instruct 1383; Gemma 4 31B 1452, 26B A4B 1441. Gaps: no qwen2.5-7b/14b/32b-instruct, no qwen3-8b/14b.",
 "Yes — multiple downloadable snapshot archives",
 "Elo is relative and non-stationary: scores drift with the model pool and voter mix; use dated snapshots, not the current table, for time series. Style-control variants added mid-2024."],
["2 — best size coverage (to Mar 2025)", "HF Open LLM Leaderboard v1 + v2 (archived)", "v1 raw: HF dataset open-llm-leaderboard-old/results; v2 table: HF dataset open-llm-leaderboard/contents (~4.6k models, Parquet); archived UIs: HF collection OpenEvals",
 "v1 (2023 - Jun 2024): ARC, HellaSwag, MMLU, TruthfulQA, Winogrande, GSM8K. v2 (Jun 2024 - Mar 2025): IFEval, BBH, MATH-Lvl5, GPQA, MUSR, MMLU-Pro",
 "v1: every Qwen1.5-Chat size 0.5B-32B, gemma-2b/7b-it, gemma-1.1. v2: full Qwen2/2.5 and Gemma 2 lines (e.g., Qwen2.5-7B-Instruct avg 35.20, gemma-2-9b-it avg 32.07). Zero Gemma 3 or Qwen3 rows (frozen Mar 2025).",
 "Yes — fully public HF datasets",
 "Uniform harness within each era, but v1 and v2 suites are disjoint (hard break Jun 2024) and everything ends Mar 2025."],
["3 — best for 2024-2026", "Artificial Analysis", "artificialanalysis.ai/models/open-source; per-model pages persist (e.g., /models/gemma-3-12b)",
 "Composite Intelligence Index (v4.1.1: GDPval-AA, tau3-Banking, Terminal-Bench, SciCode, HLE, GPQA-D, CritPt, AA-Omniscience, AA-LCR) + price/speed",
 "~746 models; persistent pages for deprecated models, re-scored on current index (gemma-3-12b Index 6; Qwen3-8B Index 5). Thin before 2024: no Gemma 1, little Qwen1.5.",
 "No bulk dump; website/API; third-party mirrors exist",
 "Index redefined repeatedly (v1→v4): published composite values are unstable over time — use per-benchmark components (e.g., GPQA) for time series."],
["Supplement — capability detail", "LiveBench", "livebench.ai; github.com/LiveBench/LiveBench; HF datasets livebench/*",
 "6 categories, 18 tasks, objective ground-truth scoring",
 "195+ models: qwen2.5-7b-instruct-turbo, qwen2.5-coder-32b, qwq-32b, gemma-2-27b-it, gemma-3-4b/12b/27b-it, Qwen3 series. Nothing before Jun 2024 launch.",
 "Yes — HF datasets + site CSVs",
 "Questions rotate every few months by design; scores only comparable within a question release."],
["Supplement — writing quality", "EQ-Bench / Creative Writing (eqbench.com)", "eqbench.com; data hardcoded in github.com/EQ-bench/EQ-bench-site JS",
 "LLM-judged emotional intelligence, creative/longform writing, slop metrics",
 "~130 models on Creative Writing v3: gemma-2-9b-it, gemma-3 and gemma-4 series, qwq-32b, larger Qwen3; small Qwen3 dense spotty. Legacy boards cover 2023-24 era.",
 "Yes in practice (CSV in site JS on GitHub)",
 "Judge models change per version; versions not cross-comparable; single maintainer, subjective domain."],
["Niche — code editing", "Aider polyglot leaderboard", "aider.chat/docs/leaderboards/ (results as YAML in repo)",
 "Code-editing success on 225 Exercism problems",
 "Qwen3-32B 40.0%, Qwen2.5-Coder-32B 16.4%; no Gemma models.",
 "Yes — YAML in GitHub repo",
 "Benchmark swapped (code-editing → polyglot) Dec 2024, breaking continuity."],
["Niche — frozen 2024 snapshot", "MixEval", "mixeval.github.io",
 "Ground-truth benchmark mixture correlated with Arena",
 "Still lists Qwen1.5-7B-Chat and Gemma-1.1-7B-IT — useful for the 2024 small-model era.",
 "Yes",
 "Frozen since ~late 2024; dynamic updates stopped."],
["Niche — deep small-model coverage", "Dubesor bench table", "dubesor.de/benchtable",
 "One-person private 4-category suite",
 "Unusually deep small/quantized Qwen + Gemma coverage 2024-25.",
 "Viewable; marked [ARCHIVED]",
 "Private questions, single evaluator; now archived."],
["Low relevance for this project", "Others: SEAL (Scale AI), BigCode leaderboard, OpenCompass, HELM", "labs.scale.com/leaderboard; HF space bigcode/bigcode-models-leaderboard; rank.opencompass.org.cn; crfm.stanford.edu/helm",
 "Various",
 "SEAL: frontier-only, no small open models. BigCode: code-only, dormant. OpenCompass: strong Qwen coverage but quarterly benchmark revisions and poor archiving. HELM: most transparent methodology, fully downloadable, but lags on new small models.",
 "Varies (HELM: excellent; SEAL: none)",
 "See coverage column."],
]

FONT = "Arial"
ws.append(headers)
for r in rows:
    ws.append(r)

header_font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
header_fill = PatternFill("solid", fgColor="1F3864")
body_font = Font(name=FONT, size=10)
wrap = Alignment(wrap_text=True, vertical="top")
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

widths = [22, 26, 42, 36, 60, 22, 44]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(vertical="center", wrap_text=True)
for row in ws.iter_rows(min_row=2, max_row=len(rows) + 1, max_col=len(widths)):
    for c in row:
        c.font = body_font
        c.border = border
        c.alignment = wrap
ws.freeze_panes = "A2"

wb.save("./qwen_gemma_benchmarks.xlsx")
print("third-party sheet added; sheets:", wb.sheetnames)
